# Программа для заполнения основной надписи КД в формате PDF


import io
import sys
from openpyxl import load_workbook
from pathlib import Path
from pypdf import PdfReader, PdfWriter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
import re


def document_is_bom(name: Path) -> bool:
    # Проверка на перечень электрических элементов
    is_electric_list = re.search(r'[\d\s+]ПЭ', name.stem) is not None
    # Проверка на таблицу соединений
    is_electric_table = re.search(r'[\d\s+]ТЭ', name.stem) is not None
    # Проверка на перечень гидравлических элементов
    is_hydraulic_list = re.search(r'[\d\s+]ПГ' ,name.stem) is not None
    # Проверка на спецификацию
    part_number = name.stem.split()[0]
    is_bom = name.stem.count('СБ') == 0 and part_number.endswith('0')
    return (is_electric_list or is_electric_table or is_hydraulic_list or is_bom)


def document_is_scheme(name: Path) -> bool:
    # Проверка на электрическую схему
    is_electric_scheme = False
    for doc_code in ['Э3', 'Э4', 'Э7']:
        if re.search(fr'[\d\s+]{doc_code}', name.stem) is not None:
            is_electric_scheme = True
    # Проверка на гидравлическую схему
    is_hydraulic_scheme = False
    for doc_code in ['Г3', 'Г4']:
        if re.search(fr'[\d\s+]{doc_code}', name.stem) is not None:
            is_hydraulic_scheme = True
    # Проверка на пневматическую схему
    is_pneumatic_scheme = False
    for doc_code in ['П3', 'П4']:
        if re.search(fr'[\d\s+]{doc_code}', name.stem) is not None:
            is_pneumatic_scheme = True
    return (is_electric_scheme or is_hydraulic_scheme or is_pneumatic_scheme)


def format_cell_value(value) -> str:
    return str(value).replace('None', '')


# Выбор файлов с расширением pdf
p = Path('.').glob('*.pdf')
file_names = sorted([name for name in p if name.is_file()])

# Выход из программы при отсутствии файлов pdf
if len(file_names) == 0:
    print('\nНе найдены файлы pdf. Завершение работы программы')
    sys.exit()

# Выход из программы при отсутствии файла с настройками
settings_path = Path('!Настройки.xlsx')
if not (settings_path.exists() and settings_path.is_file()):
    print(f'\nНе найден файл {settings_path.name}. Завершение работы программы')
    sys.exit()

# Список файлов с подписями
sign_list = [
'!Подпись_1_разработал.jpg',
'!Подпись_2_проверил.jpg',
'!Подпись_3_т.контр.jpg',
'!Подпись_4_н.контр.jpg',
'!Подпись_5_утвердил.jpg',
'!Подпись_6_согласов.jpg'
]

# Выход из программы при отсутствии файла с подписью
for name in sign_list:
    p = Path(name)
    if not (p.exists() and p.is_file()):
        print(f'\nНе найден файл {p.name}. Завершение работы программы')
        sys.exit()

# Проверка наличия папки для измененных документов
output_folder_name = 'КД_подписано'
Path(output_folder_name).mkdir(exist_ok=True)

# Получение настроек
wb = load_workbook(settings_path)
ws = wb.active

# Фамилии
developer_last_name = format_cell_value(ws['B2'].value)
check_last_name = format_cell_value(ws['B3'].value)
technology_last_name = format_cell_value(ws['B4'].value)
norm_last_name = format_cell_value(ws['B5'].value)
agreed_last_name = format_cell_value(ws['B6'].value)
approve_last_name = format_cell_value(ws['B7'].value)

# Даты
developer_date = format_cell_value(ws['C2'].value)
check_date = format_cell_value(ws['C3'].value)
technology_date = format_cell_value(ws['C4'].value)
norm_date = format_cell_value(ws['C5'].value)
agreed_date = format_cell_value(ws['C6'].value)
approve_date = format_cell_value(ws['C7'].value)

# Статус "Поставить подпись"
developer_sign_status = format_cell_value(ws['D2'].value)
check_sign_status = format_cell_value(ws['D3'].value)
technology_sign_status = format_cell_value(ws['D4'].value)
norm_sign_status = format_cell_value(ws['D5'].value)
agreed_sign_status = format_cell_value(ws['D6'].value)
approve_sign_status = format_cell_value(ws['D7'].value)

# Данные об извещении
number_of_change = format_cell_value(ws['B9'].value)
type_of_change = format_cell_value(ws['B10'].value)
document_change_number = format_cell_value(ws['B11'].value)
change_sign_status = format_cell_value(ws['B12'].value)
change_date = format_cell_value(ws['B13'].value)

# Данные о литере
litera_1 = format_cell_value(ws['B16'].value)
litera_2 = format_cell_value(ws['B17'].value)
litera_3 = format_cell_value(ws['B18'].value)

# Добавить суффикс к имени файла
add_suffix_to_filename = format_cell_value(ws['D20'].value)

wb.close()

# Список данных об извещении
document_change_data = [
number_of_change,
type_of_change,
document_change_number,
change_sign_status,
change_date
]

# Список данных о фамилиях
last_name_data = [
developer_last_name,
check_last_name,
technology_last_name,
norm_last_name,
approve_last_name,
agreed_last_name
]

# Список данных о подписях
sign_status_data = [
developer_sign_status,
check_sign_status,
technology_sign_status,
norm_sign_status,
approve_sign_status,
agreed_sign_status
]

# Список данных с датами подписей
date_data = [
developer_date,
check_date,
technology_date,
norm_date,
approve_date,
agreed_date
]

# Список данных о литере
litera_data = [
litera_1,
litera_2,
litera_3
]

# ИЗВЕЩЕНИЕ
# Координаты текстовых объектов по горизонтали от правого края спецификации (в мм)
x_change_bom = [
188, # Номер изменения
181, # Вид изменения
170, # Номер извещения
149, # Подпись
135.5, # Дата
]

# Координаты текстовых объектов по вертикали от нижнего края спецификации (в мм)
y_change_bom = [
38, # Номер изменения
38, # Вид изменения
38, # Номер извещения
37, # Подпись
38, # Дата
]

# Координаты текстовых объектов по горизонтали от правого края чертежа (в мм)
x_change_drawing = [
187, # Номер изменения
179, # Вид изменения
170, # Номер извещения
149, # Подпись
135, # Дата
]

# Координаты текстовых объектов по вертикали от нижнего края чертежа (в мм)
y_change_drawing = [
42, # Номер изменения
42, # Вид изменения
42, # Номер извещения
41, # Подпись
42, # Дата
]

# ФАМИЛИИ
# Координаты текстовых объектов по горизонтали от правого края спецификации (в мм)
x_last_name_bom = [
172, # Разработал
172, # Проверил
0,
190, # Согласовано
172, # Нормоконтроль
172 # Утвердил
]

# Координаты текстовых объектов по вертикали от нижнего края спецификации (в мм)
y_last_name_bom = [
26, # Разработал
21, # Проверил
0,
16, # Согласовано
11, # Нормоконтроль
6 # Утвердил
]

# Координаты текстовых объектов по горизонтали от правого края чертежа (в мм)
x_last_name_drawing = [
172, # Разработал
172, # Проверил
172, # Техконтроль
190, # Согласовано
172, # Нормоконтроль
172 # Утвердил
]

# Координаты текстовых объектов по вертикали от нижнего края чертежа (в мм)
y_last_name_drawing = [
31, # Разработал
26, # Проверил
21, # Техконтроль
16, # Согласовано
11, # Нормоконтроль
6 # Утвердил
]

# ПОДПИСИ
# Координаты текстовых объектов по горизонтали от правого края спецификации (в мм)
x_sign_bom = [
149, # Разработал
149, # Проверил
0,
149, # Согласовано
149, # Нормоконтроль
149 # Утвердил
]

# Координаты текстовых объектов по вертикали от нижнего края спецификации (в мм)
y_sign_bom = [
26, # Разработал
21, # Проверил
0,
16, # Согласовано
11, # Нормоконтроль
6 # Утвердил
]

# Координаты текстовых объектов по горизонтали от правого края чертежа (в мм)
x_sign_drawing = [
149, # Разработал
149, # Проверил
149, # Техконтроль
149, # Согласовано
149, # Нормоконтроль
149 # Утвердил
]

# Координаты текстовых объектов по вертикали от нижнего края чертежа (в мм)
y_sign_drawing = [
31, # Разработал
26, # Проверил
21, # Техконтроль
16, # Согласовано
11, # Нормоконтроль
6 # Утвердил
]

# ДАТЫ
# Координаты текстовых объектов по горизонтали от правого края спецификации (в мм)
x_date_bom = [
134.5, # Разработал
134.5, # Проверил
0,
134.5, # Согласовано
134.5, # Нормоконтроль
134.5 # Утвердил
]

# Координаты текстовых объектов по вертикали от нижнего края спецификации (в мм)
y_date_bom = [
26, # Разработал
21, # Проверил
0,
16, # Согласовано
11, # Нормоконтроль
6 # Утвердил
]

# Координаты текстовых объектов по горизонтали от правого края чертежа (в мм)
x_date_drawing = [
135, # Разработал
135, # Проверил
135, # Техконтроль
135, # Согласовано
135, # Нормоконтроль
135 # Утвердил
]

# Координаты текстовых объектов по вертикали от нижнего края чертежа (в мм)
y_date_drawing = [
31.5, # Разработал
26.5, # Проверил
21.5, # Техконтроль
16.5, # Согласовано
11.5, # Нормоконтроль
6.5 # Утвердил
]

# ЛИТЕРА
# Координаты текстовых объектов по горизонтали от правого края спецификации (в мм)
x_litera_bom = [
56, # Литера 1
51, # Литера 2
46 # Литера 3
]

# Координаты текстовых объектов по вертикали от нижнего края спецификации (в мм)
y_litera_bom = [
22.5, # Литера 1
22.5, # Литера 2
22.5 # Литера 3
]

# Координаты текстовых объектов по горизонтали от правого края чертежа (в мм)
x_litera_drawing = [
54.5, # Литера 1
49.5, # Литера 2
44.5 # Литера 3
]

# Координаты текстовых объектов по вертикали от нижнего края чертежа (в мм)
y_litera_drawing = [
31, # Литера 1
31, # Литера 2
31 # Литера 3
]

# Размеры подписи
sign_width = 35
sign_height = 11

# Добавление шрифта ГОСТ
pdfmetrics.registerFont(TTFont('GOST', '!GOST type A Italic.ttf'))

convert_mm_to_pt = 72 / 25.4
for name in file_names:
    # Получение размеров первой страницы
    in_stream = open(name, 'rb')
    existing_pdf = PdfReader(in_stream)
    first_page = existing_pdf.pages[0]
    pages_count = len(existing_pdf.pages)
    page_width = round(first_page.mediabox.width)
    page_height = round(first_page.mediabox.height)
    # Создание пустой PDF-страницы с текстовыми объектами
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=(page_width, page_height))
    # Определение типа документа - спецификация, схема или чертеж
    doc_is_bom = document_is_bom(name)
    doc_is_scheme = document_is_scheme(name)

    # ЗАПОЛНЕНИЕ ДАННЫХ ОБ ИЗВЕЩЕНИИ
    can.setFont('GOST', 8)
    if doc_is_bom and not doc_is_scheme:
        x_coordinates = x_change_bom
        y_coordinates = y_change_bom
    else:
        x_coordinates = x_change_drawing
        y_coordinates = y_change_drawing
    for index in range(0, len(x_coordinates)):
        x_coordinate = page_width - x_coordinates[index] * convert_mm_to_pt
        y_coordinate = y_coordinates[index] * convert_mm_to_pt
        # Уменьшение шрифта для даты
        if index == 4:
            can.setFont('GOST', 6)
        if index == 3 and change_sign_status == 'да':
            can.drawImage(sign_list[0], x_coordinate, y_coordinate, sign_width, sign_height)
        elif index == 3 and change_sign_status != 'да':
            continue
        else:
            if document_change_data[index] == '':
                continue
            can.drawString(x_coordinate, y_coordinate, document_change_data[index])

    # ЗАПОЛНЕНИЕ ФАМИЛИЙ
    can.setFont('GOST', 14)
    if doc_is_bom and not doc_is_scheme:
        x_coordinates = x_last_name_bom
        y_coordinates = y_last_name_bom
    else:
        x_coordinates = x_last_name_drawing
        y_coordinates = y_last_name_drawing
    for index in range(0, len(x_coordinates)):
        if sign_status_data[index] != 'да':
            continue
        # Пропуск фамилии технолога в спецификацях
        if doc_is_bom and index == 2:
            continue
        # Пропуск фамилии технолога в схемах
        if doc_is_scheme and index == 2:
            continue
        if last_name_data[index] == '':
            continue
        x_coordinate = page_width - x_coordinates[index] * convert_mm_to_pt
        y_coordinate = y_coordinates[index] * convert_mm_to_pt
        can.drawString(x_coordinate, y_coordinate, last_name_data[index])

    # ПРОСТАНОВКА ПОДПИСЕЙ
    if doc_is_bom and not doc_is_scheme:
        x_coordinates = x_sign_bom
        y_coordinates = y_sign_bom
    else:
        x_coordinates = x_sign_drawing
        y_coordinates = y_sign_drawing
    for index in range(0, len(x_coordinates)):
        if sign_status_data[index] != 'да':
            continue
        # Пропуск фамилии технолога в спецификацях
        if doc_is_bom and index == 2:
            continue
        # Пропуск фамилии технолога в схемах
        if doc_is_scheme and index == 2:
            continue
        x_coordinate = page_width - x_coordinates[index] * convert_mm_to_pt
        y_coordinate = y_coordinates[index] * convert_mm_to_pt
        can.drawImage(sign_list[index], x_coordinate, y_coordinate, sign_width, sign_height)

    # ЗАПОЛНЕНИЕ ДАТ
    can.setFont('GOST', 7)
    if doc_is_bom and not doc_is_scheme:
        x_coordinates = x_date_bom
        y_coordinates = y_date_bom
    else:
        x_coordinates = x_date_drawing
        y_coordinates = y_date_drawing
    for index in range(0, len(x_coordinates)):
        if sign_status_data[index] != 'да':
            continue
        # Пропуск фамилии технолога в спецификацях
        if doc_is_bom and index == 2:
            continue
        # Пропуск фамилии технолога в схемах
        if doc_is_scheme and index == 2:
            continue
        if date_data[index] == '':
            continue
        x_coordinate = page_width - x_coordinates[index] * convert_mm_to_pt
        y_coordinate = y_coordinates[index] * convert_mm_to_pt
        can.drawString(x_coordinate, y_coordinate, date_data[index])

    # ЗАПОЛНЕНИЕ ЛИТЕРЫ
    can.setFont('GOST', 10)
    if doc_is_bom and not doc_is_scheme:
        x_coordinates = x_litera_bom
        y_coordinates = y_litera_bom
    else:
        x_coordinates = x_litera_drawing
        y_coordinates = y_litera_drawing
    for index in range(0, len(x_coordinates)):
        if litera_data[index] == '':
            continue
        x_coordinate = page_width - x_coordinates[index] * convert_mm_to_pt
        y_coordinate = y_coordinates[index] * convert_mm_to_pt
        can.drawString(x_coordinate, y_coordinate, litera_data[index])

    can.save()
    packet.seek(0)
    # Объединение существующего документа с текстовыми объектами
    new_pdf = PdfReader(packet)
    first_page.merge_page(new_pdf.pages[0])
    # Сохранение измененного документа
    output_pdf = PdfWriter()
    output_pdf.add_page(first_page)
    for page_index in range(1, pages_count):
        output_pdf.add_page(existing_pdf.pages[page_index])
    # Добавление в имя измененного документа его статуса
    # _п - подписано проверяющим
    # _н - подписано нормоконтролером
    # _у - подписано утверждающим
    # _с - подписано согласующим
    file_extension = name.suffix
    file_name_suffix = ''
    output_file_name = Path(name)
    if add_suffix_to_filename == 'да':
        for status_suffix in ['_п', '_н', '_у', '_с']:
            output_file_name = Path(output_file_name.name.replace(f'{status_suffix}{file_extension}', f'{file_extension}'))
        if check_sign_status == 'да':
            file_name_suffix = '_п'
        if norm_sign_status == 'да':
            file_name_suffix = '_н'
        if approve_sign_status == 'да':
            file_name_suffix = '_у'
        if agreed_sign_status == 'да':
            file_name_suffix = '_с'
    output_file_name = Path(f'{output_file_name.stem}{file_name_suffix}{file_extension}')
    output_path = Path(output_folder_name / output_file_name)
    with open(output_path, 'wb') as out_stream:
        output_pdf.write(out_stream)
    in_stream.close()
    print(f'\nДобавлен файл {output_path.name}')

print(f'\nИзмененные документы сохранены в папку "{output_folder_name}"')
